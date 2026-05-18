import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

sheet1 = pd.read_excel('Sheet1_Individual_Marks.xlsx')
sheet2 = pd.read_excel('Sheet2_Student_Totals.xlsx')

sheet1['Calculated_Total'] = (
    sheet1['Math (out of 100)'] +
    sheet1['Science (out of 100)'] +
    sheet1['English (out of 100)']
)

merged = pd.merge(
    sheet1[['Sr. No.', 'Student Name', 'Math (out of 100)',
            'Science (out of 100)', 'English (out of 100)', 'Calculated_Total']],
    sheet2[['Sr. No.', 'Total Marks (out of 300)']],
    on='Sr. No.'
)

merged['Status'] = merged.apply(
    lambda row: 'CORRECT' if row['Calculated_Total'] == row['Total Marks (out of 300)']
                else 'WRONG',
    axis=1
)
merged['Difference'] = merged['Total Marks (out of 300)'] - merged['Calculated_Total']

correct_count = (merged['Status'] == 'CORRECT').sum()
wrong_count   = (merged['Status'] == 'WRONG').sum()

print("TOTAL VERIFICATION REPORT")
print("==========================")
print(f"Total Students  : {len(merged)}")
print(f"Correct Totals: {correct_count}")
print(f"Wrong Totals  : {wrong_count}")

wrong_df = merged[merged['Status'] == 'WRONG'][
    ['Sr. No.', 'Student Name', 'Calculated_Total',
     'Total Marks (out of 300)', 'Difference']
].rename(columns={
    'Calculated_Total': 'Correct Total',
    'Total Marks (out of 300)': 'Given Total'
})

output_file = 'Verification_Result.xlsx'
merged.to_excel(output_file, index=False)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

green_fill = PatternFill('solid', start_color='C6EFCE')
red_fill   = PatternFill('solid', start_color='FFC7CE')
green_font = Font(color='276221', bold=False)
red_font   = Font(color='9C0006', bold=False)

status_col = merged.columns.get_loc('Status') + 1 
for row in range(2, len(merged) + 2):
    cell = ws.cell(row=row, column=status_col)
    if cell.value == 'CORRECT':
        for c in range(1, len(merged.columns) + 1):
            ws.cell(row=row, column=c).fill = green_fill
        cell.font = green_font
    else:
        for c in range(1, len(merged.columns) + 1):
            ws.cell(row=row, column=c).fill = red_fill
        cell.font = red_font

wb.save(output_file)

